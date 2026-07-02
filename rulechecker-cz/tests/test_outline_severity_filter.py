import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_parser import IssueRecord
from formatter import write_output_excel


def _record(severity_cz: str, severity_en: str, identifier: str, source_file: str = "") -> IssueRecord:
    return IssueRecord(
        rc=3002,
        severity_cz=severity_cz,
        severity_en=severity_en,
        title_cz="Kontrola geometrie propojení",
        title_en="Connection geometry check",
        explanation_cz="CZ",
        explanation_en="EN",
        object_type_cz="Spoj",
        object_type_en="Connection",
        wire_number=identifier,
        affected_cz="",
        affected_en="",
        where_cz="",
        where_en="",
        recommendation_cz="",
        recommendation_en="",
        harness_name="TAB019708C_LTGS_KSK_RL",
        source_file=source_file,
        source_sheet="1_LIN-Bus_Leitungsvorgabe",
        source_row=12,
        history_excel="[Historie](file:///C:/temp/history.xlsx)",
    )


class TestOutlineSeverityFilter(unittest.TestCase):
    def test_mixed_rc_is_split_into_filterable_severity_groups(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            out_path = Path(tmp_dir) / "out.xlsx"
            write_output_excel(
                out_path,
                [
                    _record("Kritické", "Critical", "A", str(out_path)),
                    _record("Nekritické", "Non-critical", "B", str(out_path)),
                ],
            )

            ws = load_workbook(out_path)["CZ_RC_Svazek"]
            header = [cell.value for cell in ws[1]]
            severity_col = header.index("Závažnost") + 1
            rc_summary_rows = [row for row in range(2, ws.max_row + 1) if ws.row_dimensions[row].outlineLevel == 0]
            severities = [ws.cell(row, severity_col).value for row in rc_summary_rows]

            self.assertEqual(header[0], "RC chyba")
            self.assertNotIn("Počet záznamů", header)
            self.assertEqual(severities, ["Kritické", "Nekritické"])
            self.assertNotIn("Mix", severities)

            first_harness_row = 3
            harness_col = header.index("Název svazku") + 1
            self.assertIsNone(ws.cell(first_harness_row, 1).value)
            self.assertEqual(ws.cell(first_harness_row, harness_col).value, "TAB019708C_LTGS_KSK_RL")

            first_detail_row = 4
            history_col = header.index("HISTORY_EXCEL") + 1
            self.assertEqual(ws.cell(first_detail_row, 1).value, "Otevřít chybu")
            self.assertIsNotNone(ws.cell(first_detail_row, 1).hyperlink)
            self.assertIsNotNone(ws.cell(first_detail_row, history_col).hyperlink)


if __name__ == "__main__":
    unittest.main()
