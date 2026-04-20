import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_parser import IssueRecord
from formatter import write_output_excel


class TestFormatterWireWrap(unittest.TestCase):
    def test_wire_number_column_wraps_and_row_height_expands(self):
        record = IssueRecord(
            rc=110,
            severity_cz="Kritické",
            severity_en="Critical",
            title_cz="Kontrola délky speciálního vedení",
            title_en="Special cable length consistency",
            explanation_cz="CZ",
            explanation_en="EN",
            object_type_cz="Speciální drát",
            object_type_en="Special cable",
            wire_number="230003\n230004",
            affected_cz="Sonderleitung = A",
            affected_en="Sonderleitung = A",
            where_cz="Leitungen = 230003 230004",
            where_en="Leitungen = 230003 230004",
            recommendation_cz="CZ rec",
            recommendation_en="EN rec",
        )

        with tempfile.TemporaryDirectory() as tmp_dir:
            out_path = Path(tmp_dir) / "out.xlsx"
            write_output_excel(out_path, [record])

            wb = load_workbook(out_path)
            ws = wb["Prehled_CZ"]
            header_index = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            wire_col = header_index["Identifikátor"]
            wire_cell = ws.cell(row=2, column=wire_col)

            self.assertEqual(wire_cell.value, "230003\n230004")
            self.assertTrue(wire_cell.alignment.wrap_text)
            self.assertGreater(ws.row_dimensions[2].height, 15)
            self.assertIn(
                ws["A2"].fill.fgColor.value,
                {"00FDE2E4", "00F8C8CD", "00F5B5BC", "00F29CA7"},
            )
            validations = list(ws.data_validations.dataValidation)
            self.assertTrue(any(v.formula1 == '"done,in progress,N/A,false"' for v in validations))


if __name__ == "__main__":
    unittest.main()
