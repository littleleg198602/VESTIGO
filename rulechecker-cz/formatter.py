from __future__ import annotations

from pathlib import Path
import re
from urllib.parse import quote

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    OUTPUT_SHEET_CZ,
    OUTPUT_SHEET_EN,
)
from excel_parser import IssueRecord

CZ_COLUMNS = [
    "Název svazku",
    "Závažnost",
    "RC",
    "Typ objektu",
    "Identifikátor",
    "Kurzname",
    "Název chyby",
    "Vysvětlení",
    "Doporučení",
    "Priority",
    "Progress",
    "Solution",
    "Notes",
    "HISTORY_EXCEL",
    "HISTORY_MAIL",
]
EN_COLUMNS = [
    "Harness name",
    "Severity",
    "RC",
    "Object type",
    "Identifier",
    "Kurzname",
    "Error title",
    "Explanation",
    "Recommendation",
    "Priority",
    "Progress",
    "Solution",
    "Notes",
    "HISTORY_EXCEL",
    "HISTORY_MAIL",
]


def _history_link_columns(history_note: str) -> dict[str, str]:
    lines = [line.strip() for line in history_note.splitlines() if line.strip()]
    out = {}
    for idx in range(5):
        out[f"HISTORY_LINK_{idx + 1}"] = lines[idx] if idx < len(lines) else ""
    return out

def _legacy_priority(severity_en: str) -> str:
    return "Not OK" if severity_en == "Critical" else "Warning"


def _default_progress(severity_en: str) -> str:
    return "in progress" if severity_en == "Critical" else "done"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CRITICAL_FILLS = [
    PatternFill("solid", fgColor="FDE2E4"),
    PatternFill("solid", fgColor="F8C8CD"),
    PatternFill("solid", fgColor="F5B5BC"),
    PatternFill("solid", fgColor="F29CA7"),
]
NON_CRITICAL_FILLS = [
    PatternFill("solid", fgColor="FFF9DB"),
    PatternFill("solid", fgColor="FFF3BF"),
    PatternFill("solid", fgColor="FFEC99"),
    PatternFill("solid", fgColor="FFE066"),
]
ROW_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _pick_fill(severity: str, row_idx: int) -> PatternFill:
    palette = CRITICAL_FILLS if severity in {"Kritické", "Critical", "Not OK"} else NON_CRITICAL_FILLS
    return palette[row_idx % len(palette)]


def build_output_frames(records: list[IssueRecord]) -> dict[str, pd.DataFrame]:
    cz_rows = [
        {
            "Název svazku": r.harness_name,
            "Závažnost": r.severity_cz,
            "RC": r.rc,
            "Typ objektu": r.object_type_cz,
            "Identifikátor": r.wire_number,
            "Kurzname": r.kurzname,
            "Název chyby": r.title_cz,
            "Vysvětlení": r.explanation_cz,
            "Doporučení": r.recommendation_cz,
            "Priority": _legacy_priority(r.severity_en),
            "Progress": _default_progress(r.severity_en),
            "Solution": "",
            "Notes": "",
            "HISTORY_EXCEL": r.history_excel,
            "HISTORY_MAIL": r.history_mail,
        }
        for r in records
    ]
    en_rows = [
        {
            "Harness name": r.harness_name,
            "Severity": r.severity_en,
            "RC": r.rc,
            "Object type": r.object_type_en,
            "Identifier": r.wire_number,
            "Kurzname": r.kurzname,
            "Error title": r.title_en,
            "Explanation": r.explanation_en,
            "Recommendation": r.recommendation_en,
            "Priority": _legacy_priority(r.severity_en),
            "Progress": _default_progress(r.severity_en),
            "Solution": "",
            "Notes": "",
            "HISTORY_EXCEL": r.history_excel,
            "HISTORY_MAIL": r.history_mail,
        }
        for r in records
    ]

    cz_df = pd.DataFrame(cz_rows, columns=CZ_COLUMNS)
    en_df = pd.DataFrame(en_rows, columns=EN_COLUMNS)

    return {
        OUTPUT_SHEET_CZ: cz_df,
        OUTPUT_SHEET_EN: en_df,
    }


def _compose_recommendation(affected: str, where: str, recommendation: str) -> str:
    return f"{affected}; {where}; {recommendation}".strip("; ").strip()


def write_output_excel(out_path: Path, records: list[IssueRecord]) -> None:
    frames = build_output_frames(records)
    for key, df in frames.items():
        drop_cols = [col for col in df.columns if str(col).startswith("HISTORY_LINK_")]
        if drop_cols:
            frames[key] = df.drop(columns=drop_cols)
    records_by_sheet = _split_records_by_sheet(records)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet, df in frames.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

        cz_df = frames[OUTPUT_SHEET_CZ].copy()
        cz_df.to_excel(writer, sheet_name="CZ_Data", index=False)
        _write_rc_harness_outline_sheet(writer.book, cz_df)
        _write_rc_summary_sheet(writer.book, cz_df)
        _write_help_sheet(writer.book)

        for sheet in frames:
            ws = writer.book[sheet]
            _add_rc_hyperlinks(ws, records_by_sheet.get(sheet, []))
            _add_history_hyperlinks(ws, "HISTORY_EXCEL")
            _add_history_hyperlinks(ws, "HISTORY_MAIL")
            _format_sheet(ws, sheet)
            _add_priority_validation(ws)
            _add_progress_validation(ws)

        cz_data_ws = writer.book["CZ_Data"]
        _add_rc_hyperlinks(cz_data_ws, records_by_sheet.get(OUTPUT_SHEET_CZ, []))
        _add_history_hyperlinks(cz_data_ws, "HISTORY_EXCEL")
        _add_history_hyperlinks(cz_data_ws, "HISTORY_MAIL")
        _format_sheet(cz_data_ws, "CZ_Data")
        _add_priority_validation(cz_data_ws)
        _add_progress_validation(cz_data_ws)



def _sort_cz_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    sort_cols = [col for col in ["RC", "Název chyby", "Název svazku", "Typ objektu", "Identifikátor"] if col in df.columns]
    if not sort_cols:
        return df.copy()
    work = df.copy()
    work["__rc_sort"] = pd.to_numeric(work.get("RC"), errors="coerce") if "RC" in work.columns else 0
    by = ["__rc_sort"] + [col for col in sort_cols if col != "RC"]
    return work.sort_values(by=by, kind="mergesort", na_position="last").drop(columns=["__rc_sort"])


def _write_rc_harness_outline_sheet(wb, cz_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("CZ_RC_Svazek", 0)
    df = _sort_cz_dataframe(cz_df)
    columns = list(df.columns)
    detail_severity_col_name = "Závažnost detailu"
    outline_label_col_name = "RC chyba"
    outline_columns = [outline_label_col_name] + columns + ([detail_severity_col_name] if "Závažnost" in columns else [])
    ws.append(outline_columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ROW_BORDER

    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    harness_col = outline_columns.index("Název svazku") + 1 if "Název svazku" in outline_columns else None
    severity_col = outline_columns.index("Závažnost") + 1 if "Závažnost" in outline_columns else None
    detail_severity_col = outline_columns.index(detail_severity_col_name) + 1 if detail_severity_col_name in outline_columns else None

    row_idx = 2
    group_cols = ["RC"]
    if "Název chyby" in columns:
        group_cols.append("Název chyby")
    if "Závažnost" in columns:
        group_cols.append("Závažnost")
    for group_key, rc_group in df.groupby(group_cols, dropna=False, sort=False):
        if not isinstance(group_key, tuple):
            group_key = (group_key, "")
        rc_value = group_key[0]
        title_value = group_key[1] if "Název chyby" in columns and len(group_key) > 1 else ""
        severity_text = group_key[-1] if "Závažnost" in columns and len(group_key) > 1 else ""
        severity_text = "" if pd.isna(severity_text) else str(severity_text)
        rc_label = f"RC {rc_value} – {title_value}".strip(" –")
        ws.cell(row_idx, 1, rc_label)
        if severity_col:
            ws.cell(row_idx, severity_col, severity_text)
        _style_summary_row(ws, row_idx, len(outline_columns), HEADER_FILL, HEADER_FONT)
        ws.row_dimensions[row_idx].collapsed = True
        row_idx += 1

        harness_groups = rc_group.groupby("Název svazku", dropna=False, sort=False) if "Název svazku" in columns else [("", rc_group)]
        for harness, harness_group in harness_groups:
            if harness_col:
                ws.cell(row_idx, harness_col, harness)
            if severity_col:
                ws.cell(row_idx, severity_col, severity_text)
            _style_summary_row(ws, row_idx, len(outline_columns), PatternFill("solid", fgColor="D9EAD3"), Font(bold=True, color="1F1F1F"))
            ws.row_dimensions[row_idx].outlineLevel = 1
            ws.row_dimensions[row_idx].hidden = True
            ws.row_dimensions[row_idx].collapsed = True
            row_idx += 1

            for _, detail in harness_group.iterrows():
                detail_severity = str(detail.get("Závažnost", ""))
                for col_idx, col in enumerate(columns, start=2):
                    value = "" if col == "Závažnost" else detail.get(col, "")
                    ws.cell(row_idx, col_idx, value)
                if detail_severity_col:
                    ws.cell(row_idx, detail_severity_col, detail_severity)
                fill = _pick_fill(detail_severity, row_idx)
                for cell in ws[row_idx]:
                    cell.fill = fill
                    cell.border = ROW_BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row_idx].outlineLevel = 2
                ws.row_dimensions[row_idx].hidden = True
                row_idx += 1

    _apply_readable_widths(ws)
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"


def _style_summary_row(ws, row_idx: int, max_col: int, fill: PatternFill, font: Font) -> None:
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.fill = fill
        cell.font = font
        cell.border = ROW_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_rc_summary_sheet(wb, cz_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Souhrn_RC")
    cols = [c for c in ["RC", "Název chyby"] if c in cz_df.columns]
    if cols:
        summary = cz_df.groupby(cols, dropna=False).size().reset_index(name="Počet výskytů")
        if "RC" in summary.columns:
            summary["__rc_sort"] = pd.to_numeric(summary["RC"], errors="coerce")
            summary = summary.sort_values(["__rc_sort"] + [c for c in cols if c != "RC"], kind="mergesort", na_position="last").drop(columns=["__rc_sort"])
    else:
        summary = pd.DataFrame({"Počet výskytů": [len(cz_df)]})
    ws.append(list(summary.columns))
    for _, row in summary.iterrows():
        ws.append(list(row))
    _format_sheet(ws, "CZ")


def _write_help_sheet(wb) -> None:
    ws = wb.create_sheet("Navod")
    rows = [
        ["Jak používat přehled"],
        ["1. Na listu CZ_RC_Svazek klikni na plus vlevo u RC chyby."],
        ["2. Potom klikni na plus u názvu svazku."],
        ["3. Zobrazí se konkrétní konektory / objekty včetně Solution, Notes a historie."],
        ["4. Filtr ve sloupci Závažnost je určený pro souhrnné řádky; detailní řádky mají původní hodnotu ve sloupci Závažnost detailu, aby je filtr Kritické/Nekritické zbytečně nevytahoval."],
    ]
    for row in rows:
        ws.append(row)
    ws[1][0].fill = HEADER_FILL
    ws[1][0].font = HEADER_FONT
    ws.column_dimensions["A"].width = 90


def _apply_readable_widths(ws) -> None:
    preferred = {
        "RC chyba": 34, "Název svazku": 28, "RC": 12, "Typ objektu": 18, "Identifikátor": 26, "Kurzname": 20,
        "Název chyby": 34, "Vysvětlení": 48, "Doporučení": 52, "Solution": 34, "Notes": 34,
        "HISTORY_EXCEL": 42, "HISTORY_MAIL": 42, "Závažnost detailu": 18,
    }
    for idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        width = preferred.get(header, min(max(len(header) + 2, 14), 60))
        ws.column_dimensions[get_column_letter(idx)].width = width

def _split_records_by_sheet(records: list[IssueRecord]) -> dict[str, list[IssueRecord]]:
    return {
        OUTPUT_SHEET_CZ: records,
        OUTPUT_SHEET_EN: records,
    }


def _add_rc_hyperlinks(ws, sheet_records: list[IssueRecord]) -> None:
    rc_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "RC":
            rc_col_idx = idx
            break

    if rc_col_idx is None:
        return

    for row_idx, record in enumerate(sheet_records, start=2):
        if row_idx > ws.max_row:
            break
        cell = ws.cell(row=row_idx, column=rc_col_idx)
        source_uri = Path(record.source_file).resolve().as_uri()
        sheet_ref = quote(record.source_sheet, safe="")
        cell.hyperlink = f"{source_uri}#'{sheet_ref}'!A{record.source_row}"
        cell.style = "Hyperlink"


def _add_history_hyperlinks(ws, column_name: str) -> None:
    history_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == column_name:
            history_col_idx = idx
            break
    if history_col_idx is None:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=history_col_idx)
        text = str(cell.value or "")
        m = re.search(r"\[([^\]]+)\]\((file://[^)]+)\)", text)
        if not m:
            continue
        cell.value = re.sub(r"\[([^\]]+)\]\((file://[^)]+)\)", r"[\1]", text)
        cell.hyperlink = m.group(2)
        cell.style = "Hyperlink"


def _add_priority_validation(ws) -> None:
    priority_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Priority":
            priority_col_idx = idx
            break
    if priority_col_idx is None or ws.max_row < 2:
        return

    priority_col_letter = ws.cell(row=1, column=priority_col_idx).column_letter
    validation = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{priority_col_letter}2:{priority_col_letter}{ws.max_row}")


def _add_progress_validation(ws) -> None:
    progress_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Progress":
            progress_col_idx = idx
            break
    if progress_col_idx is None or ws.max_row < 2:
        return

    progress_col_letter = ws.cell(row=1, column=progress_col_idx).column_letter
    validation = DataValidation(type="list", formula1='"start,in progress,N/A,done"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{progress_col_letter}2:{progress_col_letter}{ws.max_row}")


def _format_sheet(ws, sheet_name: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 60)

    wire_col_name = "Identifikátor" if "CZ" in sheet_name else "Identifier"
    severity_col_name = "Závažnost" if "CZ" in sheet_name else "Severity"
    wire_col_idx = None
    severity_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == wire_col_name:
            wire_col_idx = idx
        if cell.value == severity_col_name:
            severity_col_idx = idx

    critical_row_idx = 0
    non_critical_row_idx = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        severity_value = ""
        if severity_col_idx:
            severity_value = str(row[severity_col_idx - 1].value or "")
        lead_value = severity_value
        if lead_value in {"Kritické", "Critical"}:
            fill = _pick_fill(lead_value, critical_row_idx)
            critical_row_idx += 1
        else:
            fill = _pick_fill(lead_value, non_critical_row_idx)
            non_critical_row_idx += 1

        max_lines = 1
        for cell in row:
            cell.fill = fill
            cell.border = ROW_BORDER
            cell.alignment = Alignment(vertical="top")
            line_count = str(cell.value or "").count("\n") + 1
            if line_count > max_lines:
                max_lines = line_count

        if wire_col_idx:
            row[wire_col_idx - 1].alignment = Alignment(wrap_text=True, vertical="top")

        if max_lines > 1:
            ws.row_dimensions[row[0].row].height = 15 * max_lines
